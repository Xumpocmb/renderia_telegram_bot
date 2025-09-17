import json
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional, List, Tuple
from openpyxl import Workbook


class MetricsCounter:
    """
    Class for counting and storing metrics of bot usage.
    Stores metrics in a JSON file.
    """
    def __init__(self, file_path: Optional[str] = None):
        """
        Initialize metrics counter.
        
        :param file_path: Path to the file where metrics will be stored
        """
        if file_path is None:
            # Default path is in the metrics directory
            base_dir = Path(__file__).parent
            file_path = os.path.join(base_dir, "bot_metrics.json")
        
        self.file_path = file_path
        self._metrics: Dict[str, int] = self._load_metrics()
    
    def _load_metrics(self) -> Dict[str, int]:
        """
        Load metrics from file if it exists.
        
        :return: Dictionary with metrics
        """
        if os.path.exists(self.file_path):
            try:
                with open(self.file_path, 'r', encoding='utf-8') as file:
                    return json.load(file)
            except (json.JSONDecodeError, IOError):
                return {}
        return {}
    
    def _save_metrics(self) -> None:
        """
        Save metrics to file.
        """
        # Ensure directory exists
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
        
        try:
            with open(self.file_path, 'w', encoding='utf-8') as file:
                json.dump(self._metrics, file, ensure_ascii=False, indent=2)
        except IOError as e:
            print(f"Error saving metrics: {e}")
    
    def increment(self, metric_name: str) -> None:
        """
        Increment the counter for the specified metric.
        
        :param metric_name: Name of the metric to increment
        """
        if metric_name in self._metrics:
            self._metrics[metric_name] += 1
        else:
            self._metrics[metric_name] = 1
        
        self._save_metrics()
    
    def get_metrics(self) -> Dict[str, int]:
        """
        Get all metrics.
        
        :return: Dictionary with all metrics
        """
        return self._metrics.copy()
    
    def export_to_xlsx(self) -> str:
        """
        Export metrics to XLSX file.
        
        :return: Path to the created XLSX file
        """
        # Create directory for reports if it doesn't exist
        reports_dir = os.path.join(os.path.dirname(self.file_path), "reports")
        os.makedirs(reports_dir, exist_ok=True)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_filename = f"metrics_report_{timestamp}.xlsx"
        xlsx_path = os.path.join(reports_dir, xlsx_filename)
        
        # Sort metrics by count (descending)
        sorted_metrics = sorted(self._metrics.items(), key=lambda x: x[1], reverse=True)
        
        # Create a workbook and select active worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Метрики бота"
        
        # Add headers
        worksheet['A1'] = 'Метрика'
        worksheet['B1'] = 'Количество вызовов'
        
        # Make headers bold
        for cell in worksheet["1:1"]:
            cell.font = cell.font.copy(bold=True)
        
        # Add data
        for row_idx, (name, count) in enumerate(sorted_metrics, start=2):
            worksheet.cell(row=row_idx, column=1, value=name)
            worksheet.cell(row=row_idx, column=2, value=count)
        
        # Auto-adjust column width
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        try:
            workbook.save(xlsx_path)
            return xlsx_path
        except IOError as e:
            print(f"Error exporting metrics to XLSX: {e}")
            return ""
    
    def get_metric(self, metric_name: str) -> int:
        """
        Get value for a specific metric.
        
        :param metric_name: Name of the metric
        :return: Value of the metric or 0 if not found
        """
        return self._metrics.get(metric_name, 0)
    
    def reset_metrics(self) -> None:
        """
        Reset all metrics to zero.
        """
        self._metrics = {}
        self._save_metrics()


# Create a singleton instance
metrics_counter = MetricsCounter()
